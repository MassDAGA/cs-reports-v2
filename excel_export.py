"""Styled Excel export for cs-reports-v2.

Builds a single workbook mirroring the dashboard: a Summary sheet of KPIs plus one sheet
per detail table. Styling helpers ported verbatim from the validated cs_cadence_report_v2.py
static-report script.
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BLUE = "1E3A5F"; LBLUE = "D6E4F0"; WHITE = "FFFFFF"
GREY = "F5F5F5"; RED = "C0392B"; AMBER = "E67E22"; GREEN = "1E7145"


def _fmt(v):
    return "N/A" if v is None or (isinstance(v, float) and pd.isna(v)) else v


def style_header(ws, row, cols, bg=BLUE, fg=WHITE):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=fg, size=11)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def write_table(ws, df, start_row, header_bg=BLUE):
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="(no cases)").font = Font(italic=True, color="888888")
        return
    style_header(ws, start_row, len(df.columns), bg=header_bg)
    for ci, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=col_name)
    for ri, row_data in enumerate(df.itertuples(index=False), start_row + 1):
        fill = PatternFill("solid", fgColor=GREY if ri % 2 == 0 else WHITE)
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")


def add_kpi(ws, row, col, label, value, color=BLUE):
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = PatternFill("solid", fgColor=color)
    lc.font = Font(bold=True, color=WHITE, size=10)
    lc.alignment = Alignment(horizontal="center", wrap_text=True)
    vc = ws.cell(row=row + 1, column=col, value=_fmt(value))
    vc.fill = PatternFill("solid", fgColor=LBLUE)
    vc.alignment = Alignment(horizontal="center")
    vc.font = Font(bold=True, size=14, color=BLUE)


def _title(ws, text, span_cols, color=BLUE):
    t = ws.cell(row=1, column=1, value=text)
    t.font = Font(bold=True, size=14, color=WHITE)
    t.fill = PatternFill("solid", fgColor=color)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"A1:{get_column_letter(span_cols)}1")
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False


def build_workbook(results: dict) -> io.BytesIO:
    """results = metrics.compute_all(...). Returns an in-memory .xlsx (BytesIO)."""
    r = results
    wb = Workbook()

    # ── Summary ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "CS Reports v2 — Summary", 8)
    ws.cell(row=2, column=1,
            value=f"Report Date: {pd.Timestamp.now().strftime('%B %d, %Y')}").font = Font(italic=True, color="888888")

    ws.cell(row=4, column=1, value="Update Timing").font = Font(bold=True, size=12, color=BLUE)
    add_kpi(ws, 5, 1, "Total Cases (Open+Closed)", r["total_cases_all"], BLUE)
    add_kpi(ws, 5, 2, "Open Cases", r["days_since_update"]["total_open"], BLUE)
    add_kpi(ws, 5, 3, "Avg Days Since Update (Open)", r["days_since_update"]["overall_avg"], GREEN)
    add_kpi(ws, 5, 4, "Avg Gap Between Updates", r["update_gaps"]["overall_avg"], GREEN)
    add_kpi(ws, 5, 5, "Avg Resp. to Installer Comment", r["resp_installer"]["overall_avg"], GREEN)
    add_kpi(ws, 5, 6, "Avg Resp. to Incoming Email", r["resp_incoming"]["overall_avg"], GREEN)

    ws.cell(row=8, column=1, value="Unanswered External Inputs").font = Font(bold=True, size=11, color=BLUE)
    add_kpi(ws, 9, 1, "Installer Comments Unanswered", r["resp_installer"]["unanswered"], AMBER)
    add_kpi(ws, 9, 2, "Incoming Emails Unanswered", r["resp_incoming"]["unanswered"], AMBER)

    ws.cell(row=12, column=1, value="Backlog & Data-Quality Flags (Open Cases)").font = Font(bold=True, size=12, color=BLUE)
    add_kpi(ws, 13, 1, "CS Queue > 20 Days (excl. L2)", r["cs_queue_over_20"]["count"], RED)
    add_kpi(ws, 13, 2, "Removed (escalated to L2)", r["cs_queue_over_20"]["removed_l2_count"], AMBER)
    add_kpi(ws, 13, 3, "Cases > 1d Without Account", r["no_account"]["count"], RED)
    add_kpi(ws, 13, 4, "Missing Contact / CS Owner", r["missing_contact_owner"]["count"], RED)
    auto_width(ws)

    # ── Detail sheets ────────────────────────────────────────────────────
    def add_table_sheet(name, title, df, header_bg=BLUE, note=None):
        s = wb.create_sheet(name)
        ncols = max(len(df.columns) if df is not None and not df.empty else 1, len(title) // 8, 1)
        _title(s, title, min(max(ncols, 3), 12), color=header_bg)
        start = 3
        if note:
            s.cell(row=2, column=1, value=note).font = Font(italic=True, color="888888", size=9)
        write_table(s, df, start, header_bg=header_bg)
        auto_width(s)

    add_table_sheet("Days Since Update by Owner",
                    "Avg Days Since Last Update — By CS Owner (Open Cases)",
                    r["days_since_update"]["per_owner"])
    add_table_sheet("Update Gaps by Owner",
                    "Avg Gap Between Internal Updates — By CS Owner (Open + Closed)",
                    r["update_gaps"]["per_owner"])
    add_table_sheet("Resp — Installer Comment",
                    "Avg Response Time to Installer Comment — By CS Owner",
                    r["resp_installer"]["per_owner"],
                    note=f"{r['resp_installer']['unanswered']} installer comment(s) had no subsequent internal update (excluded from average).")
    add_table_sheet("Resp — Incoming Email",
                    "Avg Response Time to Incoming Email — By CS Owner",
                    r["resp_incoming"]["per_owner"],
                    note=f"{r['resp_incoming']['unanswered']} incoming email(s) had no subsequent internal update (excluded from average).")
    add_table_sheet("CS Queue over 20d",
                    f"Open CS-Queue Cases > 20 Days ({r['cs_queue_over_20']['count']} cases)",
                    r["cs_queue_over_20"]["cases"], header_bg=RED,
                    note=f"{r['cs_queue_over_20']['removed_l2_count']} case(s) removed because they were escalated to L2 at some point.")
    add_table_sheet("No Account",
                    f"Open Cases > 1 Day Without an Account ({r['no_account']['count']} cases)",
                    r["no_account"]["cases"], header_bg=RED)
    add_table_sheet("Missing Contact or Owner",
                    f"Open Cases > 1 Day Missing Contact / CS Owner ({r['missing_contact_owner']['count']} cases)",
                    r["missing_contact_owner"]["cases"], header_bg=RED)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
